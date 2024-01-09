#Leer documentación de python para el filtrado de los datos teniendo en cuenta que serán tomados desde la comparación con otros archivos.
# HACER LAS CONEXIONES DE LAS DEMÁS HOJAS PARA SEGUIR HACIENDO PRUEBAS

# importamos las librerias necesarias
import pandas as pd 
import openpyxl as op
import os
import datetime as daytimne
import datetime as timeland

# Subprograma para calcular el porcentaje de un nombre
def calcular_porcentaje_coincidencia(nombre1, nombre2):
    nombre1 = nombre1.lower().split() # Converitimos en minusculas las variables y las separamos
    nombre2 = nombre2.lower().split()

    coincidencias = sum(palabra in nombre2 for palabra in nombre1) # Calculamos la councidencia de palabras
    porcentaje_coincidencia = coincidencias / max(len(nombre1), len(nombre2)) # Determinamos el porcentaje
    return porcentaje_coincidencia



if __name__ == "__main__":

    # Carpeta donde se encuentran los archivos
    path = r'\\csfiles\CSArchivos\Dir_Comercial\Cobros\Planes Transversales\Cartera_Financiada\BD_RCR\\'
    path1 = r"\\csfiles\CSArchivos\Dir_Comercial\Cobros\Acuerdos_de_Pago\Seguimiento_EnciendeTuEnergia\Financiaciones.txt"
    path2 = r'\\csfiles\CSArchivos\Dir_Comercial\Cobros\Deuda_Diaria\Deuda_Diaria_30-12-2023.txt'
    #path = p('C:\Users\melany.rodriguez\OneDrive - Air-e SAS ESP\Pruebas\\')

    # En esta área colocaremos la apertura de los archivos de tipo txt 
   
    year = 2022
    posicion_columna = 17
    filtered_lines_acuerdos = []
    
    # Apertura de la ruta para revisdar el archivo
    with open(path1, 'r', encoding='latin-1') as file:
        lineas = file.readlines()[1:]# Lectura de los datos del archivo
        for linea in lineas:
            partes = linea.strip().split("|") # División de los datos del archivo para poder procesarlos
            valor_columna = partes[posicion_columna].strip()
            if int(valor_columna) >= year:
                filtered_lines_acuerdos.append(partes)

    # Creación del archivo excel para guardar los datos tomados del txt
    column_names_acuerdos = ["NUMERACION", "TERRITORIAL", "SERVICIO", "CUENTA", "PRODUCTO", "CATEGORIA", "SUBCATEGORIA", "TARIFA", "TITULAR_CONTRATO", "NUM_ID_TITULAR", "TIPO_DOC_TITULAR",	"TELE_TITULAR",	"CORREO_TITULAR", "CANT CLIENTES", "ID_SOLICITUD", "ID_FINANCIACION", "FECHA_FINANCIACION",	"AÑO_ACU",	"PERIODO_ACU",	"MONTO_FINANCIADO",	"SALDO_FINANCIADO",	"ESTADO ACU",	"ID_PLAN", "PLAN_FINANCIACION",	"CLASIFICACION PLAN",	"PLAZO_FINANCIACION",	"NUM PAGARE",	"CI PACTADA",	"CUOTA MENSUAL",	"CUPON CI",	"TIPO_CUPON",	"FUE_PAGADO",	"PAGO CI ENERGIA",	"PAGO CI TERCEROS",	"PAGO CI TOTAL",	"FECHA_PAGO CI",	"FECHA_APLICACION CI",	"CANT CUOTAS FACT",	"VLR_CUOTAS_FACT",	"CANT CUOTAS PAGADAS",	"VLR_CUOTAS_PAGADAS", "CANT CUOTAS PDTE PAGO",	"VLR CUOTAS PDTE PAGO",	"CANT CUOTAS NO VENC PDTE PAGO",	"VLR CUOTAS NO VENC PDTE PAGO",	"CANT CUOTAS VENC PDTE PAGO",	"VLR CUOTAS VENC PDTE PAGO",	"CANT CUOTAS PDTE FACT",	"VLR CUOTAS PDTE FACT",	"VLR_PDTE_PAGO_EDAD_30",	"VLR_PDTE_PAGO_EDAD_60",	"VLR_PDTE_PAGO_EDAD_90",	"VLR_PDTE_PAGO_EDAD_180",	"VLR_PDTE_PAGO_EDAD_MAS_180",	"FECHA VCTO CUOTA",	"EDAD_MAX",	"DIAS VENCIDOS",	"CUMPLIMIENTO ACUERDO",	"NIVEL CUMPLIMIENTO",	"F_VCTO_ACU",	"VLR_FACT_MES",	"VLR_PAGO_MES",	"TITULAR_ACUERDO",	"NUM_ID_TITULAR_ACU",	"TIPO_DOC_TITULAR_ACU",	"TELE_TITU_SOLIC",	"CORREO_TITU_SOLIC",	"USUARIO_GENERADOR",	"NOMBRE USUARIO",	"OFICINA",	"PROCESO",	"GERENCIA",	"TERRITORIAL USUARIO",	"SUB-TERRITORIAL USU",	"DEPARTAMENTO",	"MUNICIPIO",	"CORREGIMIENTO",	"BARRIO",	"DIRECCION",	"VALOR SOCIAL",	"NOMBRE TRANSF",	"F_ACTUALIZACION",	"PERIODO_CIERRE",	"PLAN_RESCATE",	"CAUSAL NO VIGENCIA",	"PLAN_ESTIMADOS",	"ALIADO COMERCIAL",	"ALIADO COM RECONOCIDO"	]
    Tabla_Acuerdos = pd.DataFrame(filtered_lines_acuerdos, columns = column_names_acuerdos)
    Tabla_Acuerdos.to_excel('Financiaciones.xlsx', index=False)
    Tabla_A_Path = os.path.join(os.getcwd(), 'Financiaciones.xlsx')

    print("ESTOY AQUÍ2")

    with open(path2, 'r') as prd:
        lineas_deudas = prd.readlines()
        column_name_deudas = lineas_deudas[0].strip().split("|")
        Tabla_Deudas = pd.DataFrame([line.strip().split("|") for line in lineas_deudas[1:]], columns = column_name_deudas)
        output_excel_file_deudas = 'Deudas.xlsx'
        Tabla_Deudas.to_excel(output_excel_file_deudas, index=False)
        Tabla_D_Path = os.path.join(os.getcwd(), output_excel_file_deudas)
        print("ESTOY AQUÍ1")
    # Nombre del archivo que deseas procesar

    Tabla_Validaciones = "ID_VALIDADAS.xlsx"
  
    # Verificamos si el archivo que se está iterando es el que deseamos
    
    archivo3 = Tabla_Validaciones

    print(f"Procesando archivo: Financiaciones.xlsx ")
    print(f"Procesando archivo: Deudas.xlsx ")
    print(f"Procesando archivo: {archivo3}")

    # Abrimos el archivo
    wb1 = op.load_workbook(Tabla_A_Path)
    wb2 = op.load_workbook(Tabla_D_Path)
    wb3 = op.load_workbook(path + archivo3)
    
    # Obtenemos la hoja activa
    hoja1 = wb1['Sheet1']
    hoja2 = wb2['Sheet1']
    hoja3 = wb3['CONSOLIDADO']

    # Obtenemos el número de filas y columnas
    filas1 = hoja1.max_row
    columnas1 = hoja1.max_column

    filas2= hoja2.max_row
    columnas2 = hoja2.max_column 

    filas3 = hoja3.max_row
    columnas3 = hoja3.max_column
        
        # Imprimimos el numero de filas y Columnas para saber si las Procesa todas
        # Teniendo en cuenta que son muchos archivos y que hay que hacer los cruces de datos, buscar información relacionada para hacer correctamente los cruces
        
    print(filas1, columnas1)

        # Obtener la fila de encabezado para crear un mapeo de nombres de columna a índices
    fila_encabezado1 = next(hoja1.iter_rows(min_row=1, max_row=1, values_only=True))  # Mapeo correspondiente a los datos del la BD_ACUERDOS
    mapeo_columnas1 = {nombre_columna: indice_columna for indice_columna, nombre_columna in enumerate(fila_encabezado1, start=1)}

    fila_encabezado2 = next(hoja2.iter_rows(min_row=1, max_row=1, values_only=True)) # Mapeo correspondiente a los datos de la BD_DEUDAS
    mapeo_columnas2 = {nombre_columna: indice_columna for indice_columna, nombre_columna in enumerate(fila_encabezado2, start=1)}

    fila_encabezado3 = next(hoja3.iter_rows(min_row=1, max_row=1, values_only=True)) # Mapeo correspondiente a los datos de la BD_VALIDACIONES
    mapeo_columnas3 = {nombre_columna: indice_columna for indice_columna, nombre_columna in enumerate(fila_encabezado3, start=1)}
    print("ESTOY AQUÍ")
        
    for fila in range(2, filas1 + 1):
            
            # Acceder a las columnas por su nombre utilizando el diccionario mapeo_columnas
            # Entre los datos es necesrio dejar de lado aquellos que se encuentren vacíos 
            # Las columnas de los datos no son tan específicos en cuanto a la información total de los datos, la mayoría de los que están seleccionados es necesario revisar cuales son los que deben ser filtrados
        
            # Los siguientes registros corresponden a los datos relacionados a la BD_ACUERDOS
            NUMERACION = hoja1.cell(row=fila, column=mapeo_columnas1["NUMERACION"]).value
            TERRITORIAL = hoja1.cell(row=fila, column=mapeo_columnas1["TERRITORIAL"]).value
            SERVICIO = hoja1.cell(row=fila, column=mapeo_columnas1["SERVICIO"]).value
            CUENTA = hoja1.cell(row=fila, column=mapeo_columnas1["CUENTA"]).value # Será usado para corroborar datos con la tabla deudas
            PRODUCTO = hoja1.cell(row=fila, column=mapeo_columnas1["PRODUCTO"]).value  
            CATEGORIA = hoja1.cell(row=fila, column=mapeo_columnas1["CATEGORIA"]).value 
            SUBCATEGORIA = hoja1.cell(row=fila, column=mapeo_columnas1["SUBCATEGORIA"]).value
            TARIFA = hoja1.cell(row=fila, column=mapeo_columnas1["TARIFA"]).value
            TITULAR_CONTRATO = hoja1.cell(row=fila, column=mapeo_columnas1["TITULAR_CONTRATO"]).value # Se va a utilizar para verificar con el nombre en la validación por porcentaje, tener en cuenta
            NUM_ID_TITULAR = hoja1.cell(row=fila, column=mapeo_columnas1["NUM_ID_TITULAR"]).value # Debe tener más de 3 dígitos para poder ser utilizada
                                                                                                # Verificar si aparecen guinoes u otros datos no numéricos 
            TIPO_DOC_TITULAR = hoja1.cell(row=fila, column=mapeo_columnas1["TIPO_DOC_TITULAR"]).value
            TELE_TITULAR = hoja1.cell(row=fila, column=mapeo_columnas1["TELE_TITULAR"]).value
            CORREO_TITULAR = hoja1.cell(row=fila, column=mapeo_columnas1["CORREO_TITULAR"]).value # Filtro en caso de que venga con algún dato de no correspondencia o sea un correo inválido
            CANT_CLIENTES = hoja1.cell(row=fila, column=mapeo_columnas1["CANT CLIENTES"]).value
            ID_SOLICITUD = hoja1.cell(row=fila, column=mapeo_columnas1["ID_SOLICITUD"]).value 
            ID_FINANCIACION = hoja1.cell(row=fila, column=mapeo_columnas1["ID_FINANCIACION"]).value # Tener en cuenta que este aplica para todos aquellos que tengan acuaerdos, teniendo en cuenta de que hay la posibilidad de que haya clientes sin los mismos
            FECHA_FINANCIACION = hoja1.cell(row=fila, column=mapeo_columnas1["FECHA_FINANCIACION"]).value # Corresponde al año en el que fue realizado, las financiaciones deben ser iguales o mayores a 2022
            PERIODO_ACU = hoja1.cell(row=fila, column=mapeo_columnas1["PERIODO_ACU"]).value # Corresponde al periodo en el que fue realizado, el periodo debe ser igual o mayor a 202206
            MONTO_FINANCIADO = hoja1.cell(row=fila, column=mapeo_columnas1["MONTO_FINANCIADO"]).value
            SALDO_FINANCIADO = hoja1.cell(row=fila, column=mapeo_columnas1["SALDO_FINANCIADO"]).value
            ESTADO_ACU = hoja1.cell(row=fila, column=mapeo_columnas1["ESTADO ACU"]).value
            ID_PLAN = hoja1.cell(row=fila, column=mapeo_columnas1["ID_PLAN"]).value 
            PLAN_FINANCIACION = hoja1.cell(row=fila, column=mapeo_columnas1["PLAN_FINANCIACION"]).value # Los planes necesitan un filtro de trabajo dado que no todos los planes clasifcan en cuanto al reporte
            CLASIFICACION_PLAN = hoja1.cell(row=fila, column=mapeo_columnas1["CLASIFICACION PLAN"]).value
            PLAZO_FINANCIACION = hoja1.cell(row=fila, column=mapeo_columnas1["PLAZO_FINANCIACION"]).value # Los números que aparecen son cuotas en meses
            CUPON_CI = hoja1.cell(row=fila, column=mapeo_columnas1["CUPON CI"]).value
            TIPO_CUPON = hoja1.cell(row=fila, column=mapeo_columnas1["TIPO_CUPON"]).value
            FUE_PAGADO = hoja1.cell(row=fila, column=mapeo_columnas1["FUE_PAGADO"]).value
            FECHA_VCTO_CUOTA = hoja1.cell(row=fila, column=mapeo_columnas1["FECHA VCTO CUOTA"]).value
            EDAD_MAX = hoja1.cell(row=fila, column=mapeo_columnas1["EDAD_MAX"]).value # En este caso hay que tener en cuenta que la edad está contada en días, para contribuir al reporte negativo debe tener entre 90 y 540 días, y en el positivo debe ser menor a 90
                                                                                    # Verificar si aparecen guinoes u otros datos no numéricos
            DIAS_VENCIDOS = hoja1.cell(row=fila, column=mapeo_columnas1["DIAS VENCIDOS"]).value # Se cuenta en días y empieza desde el momento en el que se hizo facturación del recibo
            CUMPLIMIENTO_ACUERDO = hoja1.cell(row=fila, column=mapeo_columnas1["CUMPLIMIENTO ACUERDO"]).value
            NIVEL_CUMPLIMIENTO = hoja1.cell(row=fila, column=mapeo_columnas1["NIVEL CUMPLIMIENTO"]).value
            F_VCTO_ACU = hoja1.cell(row=fila, column=mapeo_columnas1["F_VCTO_ACU"]).value 
            VLR_FACT_MES = hoja1.cell(row=fila, column=mapeo_columnas1["VLR_FACT_MES"]).value
            VLR_PAGO_MES = hoja1.cell(row=fila, column=mapeo_columnas1["VLR_PAGO_MES"]).value
            TITULAR_ACUERDO = hoja1.cell(row=fila, column=mapeo_columnas1["TITULAR_ACUERDO"]).value # Se va a utilizar para comparar en porcentaje de validación quien hizo el acuerdo
            NUM_ID_TITULAR_ACU = hoja1.cell(row=fila, column=mapeo_columnas1["NUM_ID_TITULAR"]).value # Debe tener más de 3 dígitos para poder ser utilizada
                                                                                                 # Verificar si aparecen guinoes u otros datos no numéricos 
            TIPO_DOC_TITULAR_ACU = hoja1.cell(row=fila, column=mapeo_columnas1["TIPO_DOC_TITULAR_ACU"]).value
            TELE_TITU_SOLIC = hoja1.cell(row=fila, column=mapeo_columnas1["TELE_TITU_SOLIC"]).value
            CORREO_TITU_SOLIC = hoja1.cell(row=fila, column=mapeo_columnas1["CORREO_TITU_SOLIC"]).value # Filtro en caso de que venga con algún dato de no correspondencia o sea un correo inválido
            USUARIO_GENERADOR = hoja1.cell(row=fila, column=mapeo_columnas1["USUARIO_GENERADOR"]).value # En este se debe poner un filtro con el que se debe excluir a un dato es específico
            NOMBRE_USUARIO = hoja1.cell(row=fila, column=mapeo_columnas1["NOMBRE USUARIO"]).value
            OFICINA = hoja1.cell(row=fila, column=mapeo_columnas1["OFICINA"]).value # Se debe crear el filtro pertinente a los datos que conciernen a la oficina
            PROCESO = hoja1.cell(row=fila, column=mapeo_columnas1["PROCESO"]).value # Same as oficina
            GERENCIA = hoja1.cell(row=fila, column=mapeo_columnas1["GERENCIA"]).value
            TERRITORIAL_USUARIO = hoja1.cell(row=fila, column=mapeo_columnas1["TERRITORIAL USUARIO"]).value
            SUB_TERRITORIAL_USU = hoja1.cell(row=fila, column=mapeo_columnas1["SUB TERRITORIAL USU"]).value
            DEPARTAMENTO = hoja1.cell(row=fila, column=mapeo_columnas1["DEPARTAMENTO"]).value
            MUNICIPIO = hoja1.cell(row=fila, column=mapeo_columnas1["MUNICIPIO"]).value
            CORREGIMIENTO = hoja1.cell(row=fila, column=mapeo_columnas1["CORREGIMIENTO"]).value
            BARRIO = hoja1.cell(row=fila, column=mapeo_columnas1["BARRIO"]).value
            DIRECCION = hoja1.cell(row=fila, column=mapeo_columnas1["DIRECCION"]).value # Se va a utilizar para verificar si ha sido entregado una notificación previa
            VALOR_SOCIAL = hoja1.cell(row=fila, column=mapeo_columnas1["VALOR SOCIAL"]).value
            F_ACTUALIZACION = hoja1.cell(row=fila, column=mapeo_columnas1["F_ACTUALIZACION"]).value
            PERIODO_CIERRE = hoja1.cell(row=fila, column=mapeo_columnas1["PERIODO_CIERRE"]).value
            PLAN_RESCATE = hoja1.cell(row=fila, column=mapeo_columnas1["PLAN_RESCATE"]).value
            CAUSAL_NO_VIGENCIA = hoja1.cell(row=fila, column=mapeo_columnas1["CAUSAL NO VIGENCIA"]).value
            PLAN_ESTIMADOS = hoja1.cell(row=fila, column=mapeo_columnas1["PLAN_ESTIMADOS"]).value

            # Se hace la transformación al dato que sea pertinente de int a str para los datos que sean pertinentes

            NUM_ID_TITULAR = str(NUM_ID_TITULAR)
            NUM_ID_TITULAR_ACU = str(NUM_ID_TITULAR_ACU)
            EDAD_MAX = str(EDAD_MAX)
            EDAD_MAX = EDAD_MAX.replace('-', ' ')
            NUM_ID_TITULAR = NUM_ID_TITULAR.replace('-', ' ')
            NUM_ID_TITULAR_ACU = NUM_ID_TITULAR_ACU.replace('-', ' ')
            EDAD_MAX = int(EDAD_MAX)

            # Filtros de datos correspondientes a cada criterio

            # Filtro correspondiente a la oficina de trabajo
            if( OFICINA != 'Call Center' or PROCESO != 'Call Center' ) :
                # Filtro correspondiente al tipo de usuario de digitación
                if(USUARIO_GENERADOR != 'INTEGRATION'):
                    # Verificación del tipo de planes que se en cuentran para ser excluidos
                    if( PLAN_FINANCIACION.casefold().__contains__(' Multiservicios'or 'Materiales'or 'irregularidades'or 'Masivas'or 'Padres'or 'Terceros'or 'Covid' or 'Normalización' or'Contados') ):
                        filas1+1
                    else:
                        # Verificar el periodo desde que se van a tomar los datos
                        if( PERIODO_ACU >= 202206):
                            # Verifiacion de edades máximas existentes
                            if(EDAD_MAX >= 90 and EDAD_MAX <= 540 ):
                                # Limpieza de valores especiales
                                if( CORREO_TITULAR.casefold().__contains__('00' or 'notiene' or 'noaporta' or '000' or '123')):
                                    CORREO_TITULAR = ' '

                                if( CORREO_TITU_SOLIC.casefold().__contains__('00' or 'notiene' or 'noaporta' or '000' or '123')):
                                    CORREO_TITU_SOLIC = ' '

                                for fila in range(2, filas2 + 1):
                                    # Los siguientes datos hacen referencia a la BD_DEUDAS
                                    FECHA_DE = hoja2.cell(row=fila, column=mapeo_columnas2["FECHA_DE"]).value
                                    CUENTA_DEUDA = hoja2.cell(row=fila, column=mapeo_columnas2["CUENTA_DEUDA"]).value
                                    CARTERA_ENERGIA_TOTAL = hoja2.cell(row=fila, column=mapeo_columnas2["CARTERA_ENERGIA_TOTAL"]).value
                                    CANT_FACT = hoja2.cell(row=fila, column=mapeo_columnas2["CANT_FACT"]).value
                                    EDAD = hoja2.cell(row=fila, column=mapeo_columnas2["EDAD"]).value

                                    # Filtros correspondientes a esta sección
                                    if (CUENTA == CUENTA_DEUDA):
                                        today = daytimne.now()
                                        FECHA_VENCIMIENTO= today - timeland(days = EDAD)
                                        if (CANT_FACT>= 3 and CANT_FACT <= 18):
                                            if( CARTERA_ENERGIA_TOTAL >= 195000):
                                                APLICA_NOTIFICACION = 1  
                                            else: 
                                                APLICA_NOTIFICACION = 0 
                                        else:
                                            APLICA_NOTIFICACION = 0
                                        
                                        for fila in range(2, filas3 + 1):

                                            TIPO_DE_IDENTIFICACION = hoja3.cell(row=fila, column=mapeo_columnas3["Tipo-Id"]).value
                                            NUMERO_ID = hoja3.cell(row=fila, column=mapeo_columnas3["Numero-Id"]).value
                                            NOMBRE_VAL = hoja3.cell(row=fila, column=mapeo_columnas3["Nombres"]).value
                                            ESTADO = hoja3.cell(row=fila, column=mapeo_columnas3["ESTADO"]).value

                                            nombre1 = TITULAR_ACUERDO
                                            nombre2 = NOMBRE_VAL
                                            nombre3 = TITULAR_CONTRATO

                                            porcentaje1 = calcular_porcentaje_coincidencia(nombre2 , nombre1)
                                            porcentaje2 = calcular_porcentaje_coincidencia(nombre2 , nombre3) 

                                            if (porcentaje1 > porcentaje2 ):
                                                if(porcentaje1 >= 0.25):
                                                        NUMERO_DE_IDENTIFICACION = NUMERO_ID
                                                        NOMBRE_COMPLETO = NOMBRE_VAL
                                                        CORREO_ELECTRONICO = CORREO_TITU_SOLIC
                                            elif(porcentaje1 < porcentaje2):
                                                if(porcentaje1 >= 0.25):
                                                        NUMERO_DE_IDENTIFICACION = NUMERO_ID
                                                        NOMBRE_COMPLETO = NOMBRE_VAL
                                                        CORREO_ELECTRONICO = CORREO_TITULAR
                                            else:
                                                if(porcentaje1 >= 0.25):
                                                        NUMERO_DE_IDENTIFICACION = NUMERO_ID
                                                        NOMBRE_COMPLETO = NOMBRE_VAL
                                                        CORREO_ELECTRONICO = CORREO_TITULAR

                                            # primera fila que se escribe
                                            uno = f''' TIPO ID   NUMERO ID   NOMBRE COMPLETO   NUMERO DE OBLIGACION   SALDO EN MORA   FECHA DE VENCIMIENTO   DIRECCION   CIUDAD   DEPARTAMENTO   CORREO ELECTRONICO   ID FINANCIACION  NUMERACION	TERRITORIAL	SERVICIO   PRODUCTO	  CATEGORIA	  SUBCATEGORIA   TARIFA	  CANT CLIENTES	  ID_SOLICITUD   FECHA_FINANCIACION   PERIODO_ACU	MONTO_FINANCIADO	SALDO_FINANCIADO	ESTADO ACU	ID_PLAN   PLAN_FINANCIACION   CLASIFICACION PLAN   PLAZO_FINANCIACION	CUPON CI   TIPO_CUPON   FUE_PAGADO   FECHA VCTO CUOTA   EDAD_MAX   DIAS VENCIDOS   CUMPLIMIENTO ACUERDO   NIVEL CUMPLIMIENTO   F_VCTO_ACU   VLR_FACT_MES   VLR_PAGO_MES   USUARIO_GENERADOR   NOMBRE USUARIO   OFICINA   PROCESO   GERENCIA   TERRITORIAL USUARIO   SUB TERRITORIAL USU   DEPARTAMENTO   CORREGIMIENTO   BARRIO   VALOR SOCIAL   PLAN_RESCATE	CAUSAL NO VIGENCIA	PLAN_ESTIMADOS'''

                                            # segunda fila que se escribe (fila que se repite }[datos de el cliente])
                                            dos = f'''{TIPO_DE_IDENTIFICACION}{NUMERO_DE_IDENTIFICACION}{NOMBRE_COMPLETO}{CUENTA}{CARTERA_ENERGIA_TOTAL}{FECHA_VENCIMIENTO}{DIRECCION}{MUNICIPIO}{DEPARTAMENTO}{CORREO_ELECTRONICO}{ID_FINANCIACION}{APLICA_NOTIFICACION}{NUMERACION}{TERRITORIAL}{SERVICIO}{PRODUCTO}{CATEGORIA}{SUBCATEGORIA}{TARIFA}{CANT_CLIENTES}{ID_SOLICITUD}{FECHA_FINANCIACION}{PERIODO_ACU}{MONTO_FINANCIADO}{SALDO_FINANCIADO}{ESTADO_ACU}{ID_PLAN}{PLAN_FINANCIACION}{CLASIFICACION_PLAN}{PLAZO_FINANCIACION}{CUPON_CI}{TIPO_CUPON}{FUE_PAGADO}{FECHA_VCTO_CUOTA}{EDAD_MAX}{DIAS_VENCIDOS}{CUMPLIMIENTO_ACUERDO}{NIVEL_CUMPLIMIENTO}{F_VCTO_ACU}{VLR_FACT_MES}{VLR_PAGO_MES}{USUARIO_GENERADOR}{NOMBRE_USUARIO}{OFICINA}{PROCESO}{GERENCIA}{TERRITORIAL_USUARIO}{SUB_TERRITORIAL_USU}{DEPARTAMENTO}{CORREGIMIENTO}{BARRIO}{VALOR_SOCIAL}{PLAN_RESCATE}{CAUSAL_NO_VIGENCIA}{PLAN_ESTIMADOS}'''
                                
                                            # Tercera fila que se escribe al final del codigo
                                            tres = f'''{F_ACTUALIZACION}{PERIODO_CIERRE}'''
            
                                            print (fila)
            
                                            # se escribe la primera fila cuando la variable fila toma el valor #3 
                                            if fila==3:    
                                                with open(fr'C:\Users\melany.rodriguez\OneDrive - Air-e SAS ESP\Pruebas\Acuerdos.txt', "a+") as archivo_txt:
                                                    archivo_txt.write(str(uno) + "\n")

                                            # se escribe la segunda fila repetidamente, esta es la fila que guarda los datos de los clientes a reportar
                                            with open(fr'C:\Users\melany.rodriguez\OneDrive - Air-e SAS ESP\Pruebas\Acuerdos.txt', "a+") as archivo_txt:
                                                archivo_txt.write(str(dos) + "\n")
            
                                            # se escribe la tercera fila al finalizar de todo el archivo txt
                                            if fila==filas1:
                                                with open(fr'C:\Users\melany.rodriguez\OneDrive - Air-e SAS ESP\Pruebas\Acuerdos.txt', "a+") as archivo_txt:
                                                    archivo_txt.write(str(tres) + "\n")
                                    else:
                                        filas2+=1

                            else:
                                filas1+=1
                        else:
                            filas1+=1
                else:
                    filas1+=1
            else:
                filas1+=1 
# Cierre de los archivos Excel
    wb1.close()
    wb2.close()
    wb3.close()
                
        

